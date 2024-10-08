from pathlib import Path

import matplotlib as mpl
import pytest

mpl.use("Agg")

from openpyxl import load_workbook

from nyc_home_cost_calculator.rent import NYCRentalCostCalculator
from nyc_home_cost_calculator.simulate import SimulationResults


@pytest.fixture(scope="session")
def calculator() -> NYCRentalCostCalculator:
    return NYCRentalCostCalculator(
        initial_rent=3_000.0,
        lease_term=1,
        total_years=10,
        utility_cost=200.0,
        renters_insurance=300.0,
        moving_cost=2_000.0,
        mean_rent_increase_rate=0.03,
        rent_increase_volatility=0.01,
        mean_inflation_rate=0.02,
        inflation_volatility=0.005,
        broker_fee_rate=0.15,
        simulations=1_000,
        rng=None,
    )


@pytest.fixture(scope="session")
def results(calculator: NYCRentalCostCalculator) -> SimulationResults:
    return calculator.simulate()


def test_initialization(calculator: NYCRentalCostCalculator) -> None:
    assert calculator.initial_rent == 3_000.0
    assert calculator.lease_term == 1
    assert calculator.total_years == 10
    assert calculator.utility_cost == 200
    assert calculator.renters_insurance == 300
    assert calculator.moving_cost == 2_000.0
    assert calculator.mean_rent_increase_rate == 0.03
    assert calculator.rent_increase_volatility == 0.01
    assert calculator.mean_inflation_rate == 0.02
    assert calculator.inflation_volatility == 0.005
    assert calculator.broker_fee_rate == 0.15
    assert calculator.simulations == 1_000
    assert calculator.rng is not None


def test_simulate_costs_over_time(calculator: NYCRentalCostCalculator, results: SimulationResults) -> None:
    costs = results.profit_loss
    assert costs is not None
    assert len(costs) == (12 * results.total_years), costs.shape
    assert all(len(year_costs) == calculator.simulations for year_costs in costs), costs.shape


def test_get_cost_statistics(results: SimulationResults) -> None:
    stats = results.get_cost_statistics()
    expected_keys = ["mean", "median", "std_dev", "percentile_5", "percentile_95"]
    assert all(key in stats for key in expected_keys)
    assert all(isinstance(value, float) for value in stats.values())


def test_export_to_excel(results: SimulationResults, tmp_path: Path) -> None:
    filename = tmp_path / "test_export.xlsx"
    results.export_to_excel(str(filename))
    assert filename.exists()

    # Check if the file is a valid Excel file
    wb = load_workbook(filename)
    assert "Cost Summary" in wb.sheetnames

    ws = wb["Cost Summary"]
    assert ws["A1"].value == "Input Parameters"


@pytest.mark.parametrize(
    ("attribute", "expected_type"),
    [
        ("initial_rent", float),
        ("lease_term", int),
        ("utility_cost", float),
        ("renters_insurance", float),
        ("moving_cost", float),
        ("mean_rent_increase_rate", float),
        ("rent_increase_volatility", float),
        ("mean_inflation_rate", float),
        ("inflation_volatility", float),
        ("broker_fee_rate", float),
        ("simulations", int),
    ],
)
def test_attribute_types(calculator: NYCRentalCostCalculator, attribute: str, expected_type: type) -> None:
    assert isinstance(getattr(calculator, attribute), expected_type)
