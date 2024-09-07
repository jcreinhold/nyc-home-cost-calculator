"""This module contains the base class for the NYC Home/Rent Cost Calculator."""

from __future__ import annotations

import logging
from abc import ABC, abstractmethod
from dataclasses import dataclass
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
from matplotlib import ticker
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from nyc_home_cost_calculator.simulate import SimulationEngine, SimulationResults
from nyc_home_cost_calculator.utils import calculate_confidence_intervals

logger = logging.getLogger(__name__)


@dataclass
class LocScaleRV:
    """Abstract base class for generating random samples given a location and scale."""

    loc: float
    scale: float
    rng: np.random.Generator | None = None

    def __post_init__(self):
        """Initialize the random number generator."""
        self.rng = self.rng or np.random.default_rng()

    def __call__(self, shape: tuple[int, ...]) -> np.ndarray:
        """Generate a random sample from a normal distribution."""
        if self.rng is None:
            msg = "Random number generator is not initialized."
            raise ValueError(msg)
        return self.rng.normal(self.loc, self.scale, shape)


class AbstractSimulatorBase(ABC):
    """Abstract base class for all financial simulators."""

    def __init__(
        self,
        total_years: int,
        simulations: int,
        rng: np.random.Generator | None = None,
    ):
        """Initialize the simulator base.

        Args:
            total_years: The total number of years to simulate.
            simulations: The number of Monte Carlo simulations to run.
            rng: Custom random number generator. If None, use default numpy RNG.
        """
        self.total_years = total_years
        self.simulations = simulations
        self.rng = rng or np.random.default_rng()
        self.simulation_engine = SimulationEngine(total_years * 12, simulations, self.rng)

    @abstractmethod
    def _simulate_vectorized(self, months: np.ndarray) -> SimulationResults:
        """Run the vectorized simulation.

        This method should be implemented by subclasses to perform the actual
        simulation calculations.

        Args:
            months: A 2D numpy array of shape (total_months, num_simulations)
                    representing the months to simulate.

        Returns:
            A SimulationResults object containing the results of the simulation.
        """

    @abstractmethod
    def _get_input_parameters(self) -> list[tuple[str, Any]]:
        """Get the input parameters for the simulator.

        This method should be implemented by subclasses to return a list of
        tuples, where each tuple contains the name of a parameter and its value.

        Returns:
            A list of tuples, each containing a parameter name and its value.
        """

    def simulate(self) -> SimulationResults:
        """Run the simulation and return the results.

        This method sets up the simulation environment and calls the
        _simulate_vectorized method to perform the actual calculations.

        Returns:
            A SimulationResults object containing the results of the simulation.
        """
        return self.simulation_engine.run_simulation(self._simulate_vectorized)


class AbstractNYCCostCalculator(AbstractSimulatorBase):
    """Abstract base class for the NYC Home/Rent Cost Calculator."""

    def __init__(
        self,
        initial_cost: float,
        total_years: int,
        mean_inflation_rate: float,
        inflation_volatility: float,
        simulations: int,
        rng: np.random.Generator | None = None,
    ):
        """Initialize the NYC Home/Rent Cost Calculator.

        Args:
            initial_cost: The initial cost of the home or rent.
            total_years: The total number of years for the calculation.
            mean_inflation_rate: The mean inflation rate.
            inflation_volatility: The inflation volatility.
            simulations: The number of simulations to run.
            rng: The random number generator.
        """
        super().__init__(total_years, simulations, rng)
        self.initial_cost = initial_cost
        self.mean_inflation_rate = mean_inflation_rate
        self.inflation_volatility = inflation_volatility

    def get_cost_statistics(self, profit_loss: np.ndarray | None = None) -> dict[str, float]:
        """Calculate summary statistics of the simulated costs.

        This method runs the cost simulations and computes various statistical measures
        based on the final year's profit/loss values across all simulations.

        Args:
            profit_loss: A list of lists, where each inner list represents a month and
                contains the profit/loss values for each simulation at that month.

        Returns:
            A dictionary containing the following statistics:
                - mean: Average profit/loss
                - median: Median profit/loss
                - std_dev: Standard deviation of profit/loss
                - percentile_5: 5th percentile of profit/loss
                - percentile_95: 95th percentile of profit/loss
        """
        if profit_loss is None:
            profit_loss = self.simulate().profit_loss

        if profit_loss is None:
            msg = "Could not find cumulative costs."
            raise ValueError(msg)

        final_year_costs = profit_loss[-1]

        return {
            "mean": final_year_costs.mean(),
            "median": np.quantile(final_year_costs, 0.5),
            "std_dev": np.std(final_year_costs, ddof=1),
            "percentile_5": np.quantile(final_year_costs, 0.05),
            "percentile_95": np.quantile(final_year_costs, 0.95),
        }

    def plot(
        self,
        profit_loss: np.ndarray | None = None,
        figsize: tuple[int, int] = (12, 6),
        title: str = "Projected Costs Over Time",
        ylabel: str = "Cost ($)",
        label: str = "Average Cost",
    ) -> None:
        """Plot the projected profit/loss over time with confidence intervals.

        This method generates a line plot showing the average profit/loss for each month
        of the loan term, along with a 95% confidence interval. It also includes a
        break-even line for reference.

        Args:
            profit_loss: A list of lists, where each inner list represents a month and
                contains the profit/loss values for each simulation at that month.
            figsize: A tuple specifying the figure size in inches (width, height).
            title: The title of the plot.
            ylabel: The label for the y-axis.
            label: The label for the average cost line.
        """
        if profit_loss is None:
            profit_loss = self.simulate().profit_loss

        if profit_loss is None:
            msg = "Could not find cumulative costs."
            raise ValueError(msg)

        years = np.array(list(range(1, (12 * self.total_years) + 1))) / 12
        avg_costs, lower_bound, upper_bound = calculate_confidence_intervals(profit_loss)

        plt.figure(figsize=figsize)
        plt.plot(years, avg_costs, label=label)
        plt.fill_between(years, lower_bound, upper_bound, alpha=0.2, label="95% Confidence Interval")

        plt.title(title)
        plt.xlabel("Years")
        plt.ylabel(ylabel)
        plt.legend()
        plt.grid(visible=True)

        plt.gca().yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
        plt.gca().xaxis.set_major_locator(ticker.MultipleLocator(1))
        plt.gca().xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"{x:.0f}"))
        plt.gca().xaxis.set_minor_locator(ticker.MultipleLocator(0.25))

        plt.tight_layout()
        plt.show()

    def export_to_excel(self, filename: str, profit_loss: np.ndarray | None = None) -> None:  # noqa: C901, PLR0912, PLR0915
        """Export simulation results and input parameters to an Excel file.

        This method creates a formatted Excel workbook with three main sections:
        1. Input Parameters
        2. Cost Statistics
        3. Profit/Loss Over Time

        Args:
            filename: The name of the Excel file to be created.
            profit_loss: A list of lists, where each inner list represents a month and
                contains the profit/loss values for each simulation at that month.
        """
        if profit_loss is None:
            profit_loss = self.simulate().profit_loss

        if profit_loss is None:
            msg = "Could not find cumulative costs."
            raise ValueError(msg)

        wb = Workbook()
        ws = wb.active
        ws.title = "Cost Summary"

        # Define styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )
        alignment = Alignment(horizontal="center", vertical="center")

        # Input Parameters section
        ws["A1"] = "Input Parameters"
        ws["A1"].font = title_font
        ws.merge_cells("A1:B1")
        ws["A1"].alignment = alignment

        headers = ["Parameter", "Value"]
        ws.append(headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        params = self._get_input_parameters()
        for param in params:
            ws.append(param)

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Cost Statistics section
        start_row = ws.max_row + 2
        ws[f"A{start_row}"] = "Cost Statistics"
        ws[f"A{start_row}"].font = title_font
        ws.merge_cells(f"A{start_row}:B{start_row}")
        ws[f"A{start_row}"].alignment = alignment

        ws.append(["Statistic", "Value"])
        for cell in ws[start_row + 1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        cost_stats = self.get_cost_statistics(profit_loss)
        for stat, value in cost_stats.items():
            ws.append([stat.capitalize(), f"${value:,.2f}"])

        for row in ws.iter_rows(min_row=start_row + 2, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Costs Over Time section
        start_row = ws.max_row + 2
        ws[f"A{start_row}"] = "Costs Over Time"
        ws[f"A{start_row}"].font = title_font
        ws.merge_cells(f"A{start_row}:D{start_row}")
        ws[f"A{start_row}"].alignment = alignment

        time_headers = ["Month", "Average Cost", "Lower 95% CI", "Upper 95% CI"]
        ws.append(time_headers)
        for cell in ws[start_row + 1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        months = list(range(1, (12 * self.total_years) + 1))
        avg_costs, ci_lower, ci_upper = calculate_confidence_intervals(profit_loss)

        for month, avg, lower, upper in zip(months, avg_costs, ci_lower, ci_upper, strict=True):
            ws.append([month, f"${avg:,.2f}", f"${avg+lower:,.2f}", f"${avg+upper:,.2f}"])

        for row in ws.iter_rows(min_row=start_row + 2, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except (AttributeError, TypeError) as e:
                    logger.debug("Error: %s", e)
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        wb.save(filename)
        logger.info("Data exported to '%s'", filename)
