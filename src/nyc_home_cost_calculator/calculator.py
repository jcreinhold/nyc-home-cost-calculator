from __future__ import annotations

import math
import random
from typing import Dict, List, Tuple

import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy import stats


class NYCHomeCostCalculator:
    """A class to calculate and simulate the long-term costs and potential profits/losses of home ownership in NYC.

    This calculator takes into account various factors such as property appreciation,
    income changes, market volatility, taxes, and other associated costs of homeownership.
    It uses Monte Carlo simulation to project potential financial outcomes over time.
    """

    def __init__(
        self,
        home_price: float,
        down_payment: float,
        mortgage_rate: float,
        loan_term: int,
        initial_income: float,
        hoa_fee: float,
        insurance_rate: float,
        maintenance_rate: float,
        property_tax_rate: float,
        mean_appreciation_rate: float,
        appreciation_volatility: float,
        mean_inflation_rate: float,
        inflation_volatility: float,
        mean_income_change_rate: float,
        income_change_volatility: float,
        retirement_contribution_rate: float,
        purchase_closing_cost_rate: float = 0.04,
        sale_closing_cost_rate: float = 0.08,
        simulations: int = 10_000,
        rng: random.Random | None = None,
    ):
        """Initialize the NYCHomeCostCalculator with home purchase and financial parameters.

        Args:
            home_price (float): The purchase price of the home.
            down_payment (float): The initial down payment amount.
            mortgage_rate (float): Annual mortgage interest rate (as a decimal).
            loan_term (int): Length of the mortgage in years.
            initial_income (float): Initial annual income of the homeowner.
            hoa_fee (float): Monthly homeowners association fee.
            insurance_rate (float): Annual insurance rate as a fraction of home value.
            maintenance_rate (float): Annual maintenance cost as a fraction of home value.
            property_tax_rate (float): Annual property tax rate.
            mean_appreciation_rate (float): Expected annual home appreciation rate.
            appreciation_volatility (float): Volatility of the home appreciation rate.
            mean_inflation_rate (float): Expected annual inflation rate.
            inflation_volatility (float): Volatility of the inflation rate.
            mean_income_change_rate (float): Expected annual rate of income change.
            income_change_volatility (float): Volatility of annual income changes.
            retirement_contribution_rate (float): Percentage of income contributed to retirement.
            purchase_closing_cost_rate (float, optional): Closing costs for purchase as a percentage of home price. Defaults to 0.04.
            sale_closing_cost_rate (float, optional): Closing costs for sale as a percentage of sale price. Defaults to 0.08.
            simulations (int, optional): Number of Monte Carlo simulations to run. Defaults to 10,000.
            rng (random.Random | None, optional): Custom random number generator. If None, use default RNG. Defaults to None.
        """
        # Store all input parameters as instance variables
        self.home_price = home_price
        self.down_payment = down_payment
        self.mortgage_rate = mortgage_rate
        self.loan_term = loan_term
        self.initial_income = initial_income
        self.hoa_fee = hoa_fee
        self.insurance_rate = insurance_rate
        self.maintenance_rate = maintenance_rate
        self.property_tax_rate = property_tax_rate
        self.mean_appreciation_rate = mean_appreciation_rate
        self.appreciation_volatility = appreciation_volatility
        self.mean_inflation_rate = mean_inflation_rate
        self.inflation_volatility = inflation_volatility
        self.mean_income_change_rate = mean_income_change_rate
        self.income_change_volatility = income_change_volatility
        self.retirement_contribution_rate = retirement_contribution_rate
        self.simulations = simulations
        self.rng = rng if rng is not None else random.Random()
        self.purchase_closing_cost_rate = purchase_closing_cost_rate
        self.sale_closing_cost_rate = sale_closing_cost_rate

        # Define federal tax brackets (2023 rates)
        self.federal_brackets = [
            (0.0, 11000.0, 0.10),
            (11000.0, 44725.0, 0.12),
            (44725.0, 95375.0, 0.22),
            (95375.0, 182100.0, 0.24),
            (1821000, 231250.0, 0.32),
            (231250.0, 578125.0, 0.35),
            (578125.0, float("inf"), 0.37),
        ]

        # Define New York State tax brackets (2023 rates)
        self.ny_state_brackets = [
            (0.0, 8500.0, 0.04),
            (8500.0, 11700.0, 0.045),
            (11700.0, 13900.0, 0.0525),
            (13900.0, 80650.0, 0.0585),
            (80650.0, 215400.0, 0.0625),
            (215400.0, 1077550.0, 0.0685),
            (1077550.0, float("inf"), 0.0882),
        ]

        # Define New York City local tax rate (2023 rate)
        self.nyc_local_rate = 0.03876

    def calculate_tax(
        self, income: float, brackets: List[Tuple[float, float, float]]
    ) -> float:
        """Calculate the tax amount based on income and tax brackets.

        Args:
            income (float): The taxable income.
            brackets (List[Tuple[float, float, float]]): List of tax brackets.
                Each bracket is a tuple of (lower_bound, upper_bound, tax_rate).

        Returns:
            float: The calculated tax amount.
        """
        tax = 0.0
        for lower, upper, rate in brackets:
            if income > lower:
                # Calculate tax for the portion of income in this bracket
                tax += (min(income, upper) - lower) * rate
            if income <= upper:
                # Break if we've reached the applicable bracket
                break
        return tax

    def generate_random_rates(self) -> Tuple[float, float, float, float, float, float]:
        """Generate random rates for appreciation, inflation, income change, and tax adjustments.

        Returns:
            A tuple containing:
                - appreciation_rate: Random home appreciation rate
                - inflation_rate: Random inflation rate
                - income_change_rate: Random income change rate
                - federal_adjustment: Random federal tax rate adjustment
                - state_adjustment: Random state tax rate adjustment
                - local_adjustment: Random local tax rate adjustment
        """
        # Generate appreciation rate using t-distribution
        degrees_of_freedom = 3
        t_random = stats.t.rvs(
            degrees_of_freedom, random_state=self.rng.randint(1, 2**32 - 1)
        )
        appreciation_rate = (
            self.mean_appreciation_rate + self.appreciation_volatility * t_random
        )

        # Generate inflation rate using normal distribution
        inflation_rate = self.rng.gauss(
            self.mean_inflation_rate, self.inflation_volatility
        )

        # Generate income change rate using skew-normal distribution
        alpha = 2
        income_change_rate = stats.skewnorm.rvs(
            alpha,
            loc=self.mean_income_change_rate,
            scale=self.income_change_volatility,
            random_state=self.rng.randint(1, 2**32 - 1),
        )
        income_change_rate = max(income_change_rate, -0.5)  # Limit income loss to 50%

        # Generate random adjustments for tax rates
        federal_adjustment = self.rng.gauss(0, 0.01)
        state_adjustment = self.rng.gauss(0, 0.005)
        local_adjustment = self.rng.gauss(0, 0.002)

        return (
            appreciation_rate,
            inflation_rate,
            income_change_rate,
            federal_adjustment,
            state_adjustment,
            local_adjustment,
        )

    def calculate_effective_tax_rates(
        self, income: float, federal_adj: float, state_adj: float, local_adj: float
    ) -> Tuple[float, float, float]:
        """Calculate effective tax rates based on income and random adjustments.

        Args:
            income (float): The taxable income.
            federal_adj (float): Random adjustment to federal tax rate.
            state_adj (float): Random adjustment to state tax rate.
            local_adj (float): Random adjustment to local tax rate.

        Returns:
            Tuple[float, float, float]: Effective federal, state, and local tax rates.
        """
        federal_tax = self.calculate_tax(income, self.federal_brackets)
        state_tax = self.calculate_tax(income, self.ny_state_brackets)
        local_tax = income * self.nyc_local_rate

        # Calculate effective rates and apply random adjustments
        federal_rate = (federal_tax / income) + federal_adj
        state_rate = (state_tax / income) + state_adj
        local_rate = (local_tax / income) + local_adj

        # Ensure rates are non-negative
        return max(0, federal_rate), max(0, state_rate), max(0, local_rate)

    def calculate_tax_deduction(
        self,
        mortgage_interest: float,
        property_tax: float,
        federal_rate: float,
        state_rate: float,
        local_rate: float,
    ) -> float:
        """Calculate the tax deduction based on mortgage interest, property tax, and tax rates.

        Args:
            mortgage_interest (float): Annual mortgage interest paid.
            property_tax (float): Annual property tax paid.
            federal_rate (float): Effective federal tax rate.
            state_rate (float): Effective state tax rate.
            local_rate (float): Effective local tax rate.

        Returns:
            float: The calculated tax deduction amount.
        """
        standard_deduction = 13850  # 2023 standard deduction for single filers
        itemized_deduction = mortgage_interest + property_tax

        if itemized_deduction > standard_deduction:
            # Calculate the benefit of itemizing over taking the standard deduction
            return (itemized_deduction - standard_deduction) * (
                federal_rate + state_rate + local_rate
            )
        return 0  # No additional benefit from itemizing

    def calculate_monthly_payment(
        self, principal: float, annual_rate: float, months: int
    ) -> float:
        """Calculate the monthly mortgage payment.

        Args:
            principal (float): The loan principal amount.
            annual_rate (float): Annual interest rate (as a decimal).
            months (int): The total number of monthly payments.

        Returns:
            float: The calculated monthly payment amount.
        """
        monthly_rate = annual_rate / 12
        return (
            principal
            * (monthly_rate * (1 + monthly_rate) ** months)
            / ((1 + monthly_rate) ** months - 1)
        )

    def simulate_costs_over_time(self) -> List[List[float]]:
        """Simulate the costs and potential profits/losses of home ownership over time.

        This method runs Monte Carlo simulations for the entire loan term, taking into account
        various factors such as property appreciation, income changes, and market fluctuations.

        Returns:
            List[List[float]]: A list of lists, where each inner list represents a year and
            contains the profit/loss values for each simulation at that year.
        """
        # Calculate initial costs and loan details
        purchase_closing_costs = self.home_price * self.purchase_closing_cost_rate
        loan_amount = self.home_price - self.down_payment + purchase_closing_costs
        monthly_payment = self.calculate_monthly_payment(
            loan_amount, self.mortgage_rate, self.loan_term * 12
        )

        all_year_costs: List[List[float]] = [[] for _ in range(self.loan_term)]

        for _ in range(self.simulations):
            # Initialize simulation variables
            home_value = self.home_price
            remaining_balance = loan_amount
            current_hoa_fee = self.hoa_fee
            current_income = self.initial_income
            cumulative_cost = purchase_closing_costs

            for year in range(self.loan_term):
                # Generate random rates for this year
                (
                    appreciation_rate,
                    inflation_rate,
                    income_change_rate,
                    fed_adj,
                    state_adj,
                    local_adj,
                ) = self.generate_random_rates()

                # Calculate effective tax rates
                federal_rate, state_rate, local_rate = (
                    self.calculate_effective_tax_rates(
                        current_income, fed_adj, state_adj, local_adj
                    )
                )

                # Calculate monthly mortgage payments for the year
                annual_interest_paid = 0.0
                annual_principal_paid = 0.0
                for _ in range(12):
                    monthly_interest = remaining_balance * (self.mortgage_rate / 12.0)
                    monthly_principal = monthly_payment - monthly_interest

                    annual_interest_paid += monthly_interest
                    annual_principal_paid += monthly_principal
                    remaining_balance -= monthly_principal

                # Calculate annual costs
                property_tax = home_value * self.property_tax_rate
                insurance = home_value * self.insurance_rate
                maintenance = home_value * self.maintenance_rate
                hoa = current_hoa_fee * 12
                retirement_contribution = (
                    current_income * self.retirement_contribution_rate
                )

                # Calculate tax deductions
                tax_deduction = self.calculate_tax_deduction(
                    annual_interest_paid,
                    property_tax,
                    federal_rate,
                    state_rate,
                    local_rate,
                )

                # Sum up total annual cost
                annual_cost = (
                    monthly_payment * 12
                    + property_tax
                    + insurance
                    + maintenance
                    + hoa
                    + retirement_contribution
                    - tax_deduction
                )
                cumulative_cost += annual_cost

                # Calculate profit/loss for this year
                sale_closing_costs = home_value * self.sale_closing_cost_rate
                profit_loss = (
                    (home_value - self.home_price)
                    + (loan_amount - remaining_balance)
                    - cumulative_cost
                    - sale_closing_costs
                )
                all_year_costs[year].append(profit_loss)

                # Update values for next year
                home_value *= 1 + appreciation_rate
                current_hoa_fee *= 1 + inflation_rate
                current_income *= 1 + income_change_rate

        return all_year_costs

    def get_cost_statistics(self) -> Dict[str, float]:
        """Calculate summary statistics of the simulated costs.

        This method runs the cost simulations and computes various statistical measures
        based on the final year's profit/loss values across all simulations.

        Returns:
            Dict[str, float]: A dictionary containing the following statistics:
                - mean: Average profit/loss
                - median: Median profit/loss
                - std_dev: Standard deviation of profit/loss
                - percentile_5: 5th percentile of profit/loss
                - percentile_95: 95th percentile of profit/loss
        """
        costs = self.simulate_costs_over_time()
        final_year_costs = costs[-1]  # Get the profit/loss values for the final year

        # Calculate mean
        mean_cost = sum(final_year_costs) / len(final_year_costs)

        # Sort costs for percentile calculations
        sorted_costs = sorted(final_year_costs)

        return {
            "mean": mean_cost,
            "median": sorted_costs[len(final_year_costs) // 2],
            "std_dev": math.sqrt(
                sum((x - mean_cost) ** 2 for x in final_year_costs)
                / len(final_year_costs)
            ),
            "percentile_5": sorted_costs[int(len(final_year_costs) * 0.05)],
            "percentile_95": sorted_costs[int(len(final_year_costs) * 0.95)],
        }

    def plot_costs_over_time(self) -> None:
        """Plot the projected profit/loss over time with confidence intervals.

        This method generates a line plot showing the average profit/loss for each year
        of the loan term, along with a 95% confidence interval. It also includes a
        break-even line for reference.
        """
        costs_over_time = self.simulate_costs_over_time()

        years = list(range(1, self.loan_term + 1))
        avg_costs = [
            sum(year_costs) / len(year_costs) for year_costs in costs_over_time
        ]

        # Calculate 95% confidence intervals
        ci_lower = [
            stats.t.ppf(0.025, len(year_costs) - 1)
            * (stats.tstd(year_costs) / math.sqrt(len(year_costs)))
            for year_costs in costs_over_time
        ]
        ci_upper = [
            stats.t.ppf(0.975, len(year_costs) - 1)
            * (stats.tstd(year_costs) / math.sqrt(len(year_costs)))
            for year_costs in costs_over_time
        ]

        lower_bound = [avg + ci_low for avg, ci_low in zip(avg_costs, ci_lower)]
        upper_bound = [avg + ci_up for avg, ci_up in zip(avg_costs, ci_upper)]

        # Create the plot
        plt.figure(figsize=(12, 6))
        plt.plot(years, avg_costs, label="Average Profit/Loss")
        plt.fill_between(
            years, lower_bound, upper_bound, alpha=0.2, label="95% Confidence Interval"
        )
        plt.axhline(y=0, color="r", linestyle="--", label="Break-even")

        # Set plot labels and title
        plt.title("Projected Profit/Loss Over Time (Including Closing Costs)")
        plt.xlabel("Years")
        plt.ylabel("Profit/Loss ($)")
        plt.legend()
        plt.grid(True)

        # Format y-axis labels as currency
        plt.gca().yaxis.set_major_formatter(
            plt.FuncFormatter(lambda x, p: f"${x:,.0f}")
        )

        plt.tight_layout()
        plt.show()

    def export_to_excel(self, filename: str) -> None:
        """Export simulation results and input parameters to an Excel file.

        This method creates a formatted Excel workbook with three main sections:
        1. Input Parameters
        2. Cost Statistics
        3. Profit/Loss Over Time

        Args:
            filename (str): The name of the Excel file to be created.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Home Ownership Cost Summary"

        # Define styles for formatting
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Input Parameters section
        ws["A1"] = "Input Parameters"
        ws["A1"].font = Font(bold=True, size=14)
        headers = ["Parameter", "Value"]
        ws.append(headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill

        # List of input parameters
        params = [
            ("Home Price", f"${self.home_price:,.2f}"),
            ("Down Payment", f"${self.down_payment:,.2f}"),
            ("Mortgage Rate", f"{self.mortgage_rate:.2%}"),
            ("Loan Term", f"{self.loan_term} years"),
            ("Initial Annual Income", f"${self.initial_income:,.2f}"),
            ("Monthly HOA Fee", f"${self.hoa_fee:,.2f}"),
            ("Insurance Rate", f"{self.insurance_rate:.2%}"),
            ("Maintenance Rate", f"{self.maintenance_rate:.2%}"),
            ("Property Tax Rate", f"{self.property_tax_rate:.2%}"),
            ("Mean Appreciation Rate", f"{self.mean_appreciation_rate:.2%}"),
            ("Appreciation Volatility", f"{self.appreciation_volatility:.2%}"),
            ("Mean Inflation Rate", f"{self.mean_inflation_rate:.2%}"),
            ("Inflation Volatility", f"{self.inflation_volatility:.2%}"),
            ("Mean Income Change Rate", f"{self.mean_income_change_rate:.2%}"),
            ("Income Change Volatility", f"{self.income_change_volatility:.2%}"),
            (
                "Retirement Contribution Rate",
                f"{self.retirement_contribution_rate:.2%}",
            ),
            ("Number of Simulations", f"{self.simulations}"),
            ("Purchase Closing Cost Rate", f"{self.purchase_closing_cost_rate:.2%}"),
            ("Sale Closing Cost Rate", f"{self.sale_closing_cost_rate:.2%}"),
        ]

        for param in params:
            ws.append(param)

        # Apply borders to parameter cells
        for row in ws[3 : ws.max_row]:
            for cell in row:
                cell.border = border

        # Cost Statistics section
        ws["A21"] = "Cost Statistics"
        ws["A21"].font = Font(bold=True, size=14)
        stats_headers = ["Statistic", "Value"]
        ws.append(stats_headers)
        for cell in ws[22]:
            cell.font = header_font
            cell.fill = header_fill

        cost_stats = self.get_cost_statistics()
        for stat, value in cost_stats.items():
            ws.append([stat.capitalize(), f"${value:,.2f}"])

        # Apply borders to statistics cells
        for row in ws[23 : ws.max_row]:
            for cell in row:
                cell.border = border

        # Profit/Loss Over Time section
        ws["D1"] = "Profit/Loss Over Time"
        ws["D1"].font = Font(bold=True, size=14)
        time_headers = ["Year", "Average Profit/Loss", "Lower 95% CI", "Upper 95% CI"]
        ws.append(time_headers)
        for cell in ws[2][3:]:
            cell.font = header_font
            cell.fill = header_fill

        # Calculate profit/loss data
        costs_over_time = self.simulate_costs_over_time()
        years = list(range(1, self.loan_term + 1))
        avg_costs = [
            sum(year_costs) / len(year_costs) for year_costs in costs_over_time
        ]
        ci_lower = [
            stats.t.ppf(0.025, len(year_costs) - 1)
            * (stats.tstd(year_costs) / math.sqrt(len(year_costs)))
            for year_costs in costs_over_time
        ]
        ci_upper = [
            stats.t.ppf(0.975, len(year_costs) - 1)
            * (stats.tstd(year_costs) / math.sqrt(len(year_costs)))
            for year_costs in costs_over_time
        ]

        # Populate profit/loss data
        for year, avg, lower, upper in zip(years, avg_costs, ci_lower, ci_upper):
            ws.append(
                [year, f"${avg:,.2f}", f"${avg+lower:,.2f}", f"${avg+upper:,.2f}"]
            )

        # Apply borders to profit/loss cells
        for row in ws[3 : ws.max_row]:
            for cell in row[3:]:
                cell.border = border

        # Adjust column widths for better readability
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        wb.save(filename)
        print(f"Data exported to {filename}")


# Example usage
if __name__ == "__main__":
    custom_rng = random.Random(42)

    calculator = NYCHomeCostCalculator(
        home_price=1_000_000,
        down_payment=200_000,
        mortgage_rate=0.03,
        loan_term=30,
        initial_income=150_000,
        hoa_fee=500,
        insurance_rate=0.005,
        maintenance_rate=0.01,
        property_tax_rate=0.01,
        mean_appreciation_rate=0.03,
        appreciation_volatility=0.05,
        mean_inflation_rate=0.02,
        inflation_volatility=0.01,
        mean_income_change_rate=0.02,
        income_change_volatility=0.03,
        retirement_contribution_rate=0.15,
        simulations=10_000,
        rng=custom_rng,
        purchase_closing_cost_rate=0.04,
        sale_closing_cost_rate=0.08,
    )

    # Calculate and print cost statistics
    results = calculator.get_cost_statistics()
    print(f"Cost Statistics: {results}")

    # Plot costs over time
    calculator.plot_costs_over_time()

    # Export data to Excel
    calculator.export_to_excel("nyc_home_ownership_summary.xlsx")
