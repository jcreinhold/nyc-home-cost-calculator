from __future__ import annotations

import matplotlib.pyplot as plt
import numpy as np
from matplotlib import ticker
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy import stats
from tqdm.auto import tqdm


class NYCRentalCostCalculator:
    """A class to calculate and simulate the long-term costs renting in NYC.

    This calculator takes into account various factors such as moving costs,
    income changes, market volatility, taxes, and other associated costs of renting.
    It uses Monte Carlo simulation to project potential financial outcomes over time.
    """

    def __init__(
        self,
        initial_rent: float = 4_500.0,
        lease_term: float = 1.0,
        total_years: int = 30,
        utility_cost: float = 100.0,
        renters_insurance: float = 200.0,
        moving_cost: float = 2_500.0,
        mean_rent_increase_rate: float = 0.03,
        rent_increase_volatility: float = 0.02,
        mean_inflation_rate: float = 0.02,
        inflation_volatility: float = 0.01,
        broker_fee_rate: float = 0.15,
        move_probability: float = 0.1,
        rent_increase_move_threshold: float = 0.1,
        simulations: int = 5_000,
        rng: np.random.Generator | None = None,
    ):
        """Initialize the NYCRentalCostCalculator with rental and financial parameters.

        Args:
            initial_rent: The initial monthly rent. Defaults to $4,500.
            lease_term: Length of the lease term in years. Defaults to 1 year.
            total_years: Length of the time renting in years. Defaults to 30 years.
            initial_income: Initial annual income of the renter. Defaults to $150,000.
            utility_cost: Monthly utility cost. Defaults to $100 per month.
            renters_insurance: Annual renters insurance premium. Defaults to $200 per year.
            moving_cost: Initial moving costs. Defaults to $2,500.
            mean_rent_increase_rate: Expected annual rent increase rate. Defaults to 3%.
            rent_increase_volatility: Volatility of the rent increase rate. Defaults to 2%.
            mean_inflation_rate: Expected annual inflation rate. Defaults to 2% a year.
            inflation_volatility: Volatility of the inflation rate. Defaults to 1%.
            broker_fee_rate: Broker fee as a percentage of annual rent. Defaults to 15%.
            move_probability: The probability of moving after a lease is up. Defaults to 10%.
            rent_increase_move_threshold: The percentage threshold at which you move. Defaults to 10%.
            simulations: Number of Monte Carlo simulations to run. Defaults to 10,000.
            rng: Custom random number generator. If None, use default numpy RNG.
        """
        self.initial_rent = initial_rent
        self.lease_term = lease_term
        self.total_years = total_years
        self.utility_cost = utility_cost
        self.renters_insurance = renters_insurance
        self.moving_cost = moving_cost
        self.mean_rent_increase_rate = mean_rent_increase_rate
        self.rent_increase_volatility = rent_increase_volatility
        self.mean_inflation_rate = mean_inflation_rate
        self.inflation_volatility = inflation_volatility
        self.broker_fee_rate = broker_fee_rate
        self.move_probability = move_probability
        self.rent_increase_move_threshold = rent_increase_move_threshold
        self.simulations = simulations
        self.rng = rng if rng is not None else np.random.default_rng()

    def generate_random_rates(self) -> tuple[float, float]:
        """Generate random rates for rent increase, inflation, and income change.

        Returns:
            A tuple containing:
                - rent_increase_rate: Random rent increase rate
                - inflation_rate: Random inflation rate
        """
        rent_increase_rate = self.rng.normal(self.mean_rent_increase_rate, self.rent_increase_volatility)
        inflation_rate = self.rng.normal(self.mean_inflation_rate, self.inflation_volatility)
        return rent_increase_rate, inflation_rate

    def should_move(self, current_rent: float, new_rent: float) -> tuple[bool, float]:
        """Determine if a move should occur based on rent increase and random chance.

        Returns:
            A tuple (should_move: bool, new_rent: float)
        """
        rent_increase_rate = (new_rent - current_rent) / current_rent
        if rent_increase_rate > self.rent_increase_move_threshold:
            new_rent_after_move = self.calculate_new_rent_after_move(current_rent, new_rent)
            return True, new_rent_after_move
        if self.rng.random() < self.move_probability:
            return True, new_rent  # In case of a random move, keep the market rate
        return False, new_rent

    def calculate_new_rent_after_move(self, current_rent: float, new_rent: float) -> float:
        """Calculate the new rent after a move triggered by a rent increase."""
        min_new_rent = current_rent
        max_new_rent = new_rent - (self.moving_cost / 12)
        return self.rng.uniform(min_new_rent, max_new_rent)

    def simulate_costs_over_time(self) -> tuple[np.ndarray, np.ndarray]:
        """Simulate rental costs over time using Monte Carlo simulation, including random moves.

        Returns:
            A tuple containing two NumPy arrays:
            1. An array of shape (simulations, total_months) containing monthly rental costs.
            2. An array of shape (simulations, total_months) containing cumulative rental costs.
        """
        total_months = self.total_years * 12
        monthly_costs = np.zeros((total_months, self.simulations))
        cumulative_costs = np.zeros((total_months, self.simulations))

        for sim in tqdm(range(self.simulations), desc="Running simulations", unit="sim"):
            current_rent = self.initial_rent
            current_utility_cost = self.utility_cost
            moving_cost = self.moving_cost
            cumulative_cost = moving_cost + (self.initial_rent * 12.0 * self.broker_fee_rate)
            time_in_current_lease = 0.0
            accumulated_rent_increase = 0.0

            for month in range(total_months):
                if month % 12 == 0:  # Generate new rates each year
                    rent_increase_rate, inflation_rate = self.generate_random_rates()

                # Accumulate rent increase, but don't apply it yet
                accumulated_rent_increase += rent_increase_rate / 12.0  # Divide by 12 for monthly rate

                # Calculate costs for this month
                monthly_rent = current_rent
                monthly_utilities = current_utility_cost
                monthly_insurance = self.renters_insurance / 12.0
                monthly_cost = monthly_rent + monthly_utilities + monthly_insurance

                # Check if lease ends this month
                if time_in_current_lease + (1.0 / 12.0) >= self.lease_term:
                    # Apply accumulated rent increase
                    new_rent = current_rent * (1 + accumulated_rent_increase)
                    should_move, potential_new_rent = self.should_move(current_rent, new_rent)

                    if should_move:
                        monthly_cost += moving_cost + (potential_new_rent * 12.0 * self.broker_fee_rate) / 12.0
                        current_rent = potential_new_rent
                    else:
                        current_rent = new_rent

                    time_in_current_lease = 0
                    accumulated_rent_increase = 0.0
                else:
                    time_in_current_lease += 1.0 / 12.0

                cumulative_cost += monthly_cost
                monthly_costs[month, sim] = monthly_cost
                cumulative_costs[month, sim] = cumulative_cost

                # Update values for next month
                current_utility_cost *= (1.0 + inflation_rate) ** (1.0 / 12.0)
                moving_cost *= (1.0 + inflation_rate) ** (1.0 / 12.0)

        return monthly_costs, cumulative_costs

    def get_cost_statistics(self, cumulative_costs: np.ndarray | None = None) -> dict[str, float]:
        """Calculate summary statistics of the simulated rental costs.

        Args:
            cumulative_costs: A list of lists, where each inner list represents a month and
                contains the profit/loss values for each simulation at that month.

        Returns:
            A dictionary containing the following statistics:
                - mean: Average cumulative cost
                - median: Median cumulative cost
                - std_dev: Standard deviation of cumulative cost
                - percentile_5: 5th percentile of cumulative cost
                - percentile_95: 95th percentile of cumulative cost
        """
        if cumulative_costs is None:
            _, cumulative_costs = self.simulate_costs_over_time()

        final_year_costs = cumulative_costs[-1]

        return {
            "mean": final_year_costs.mean(),
            "median": np.quantile(final_year_costs, 0.5),
            "std_dev": np.std(final_year_costs, ddof=1),
            "percentile_5": np.quantile(final_year_costs, 0.05),
            "percentile_95": np.quantile(final_year_costs, 0.95),
        }

    def plot_costs_over_time(self, cumulative_costs: np.ndarray | None = None) -> None:
        """Plot the projected cumulative rental costs over time with confidence intervals.

        Args:
            cumulative_costs: A list of lists, where each inner list represents a month and
                contains the profit/loss values for each simulation at that month.
        """
        if cumulative_costs is None:
            _, cumulative_costs = self.simulate_costs_over_time()

        years = np.array(list(range(1, (12 * self.total_years) + 1))) / 12
        avg_costs = cumulative_costs.mean(axis=1)

        ci_lower = [
            stats.t.ppf(0.025, len(year_costs) - 1) * (stats.tstd(year_costs) / np.sqrt(len(year_costs)))
            for year_costs in cumulative_costs
        ]
        ci_upper = [
            stats.t.ppf(0.975, len(year_costs) - 1) * (stats.tstd(year_costs) / np.sqrt(len(year_costs)))
            for year_costs in cumulative_costs
        ]

        lower_bound = [avg + ci_low for avg, ci_low in zip(avg_costs, ci_lower, strict=True)]
        upper_bound = [avg + ci_up for avg, ci_up in zip(avg_costs, ci_upper, strict=True)]

        plt.figure(figsize=(12, 6))
        plt.plot(years, avg_costs, label="Average Cumulative Cost")
        plt.fill_between(years, lower_bound, upper_bound, alpha=0.2, label="95% Confidence Interval")

        plt.title("Projected Cumulative Rental Costs Over Time")
        plt.xlabel("Years")
        plt.ylabel("Cumulative Cost ($)")
        plt.legend()
        plt.grid(True)

        # Format axes
        plt.gca().yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, p: f"${x:,.0f}"))
        plt.gca().xaxis.set_major_locator(ticker.MultipleLocator(1))
        plt.gca().xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, p: f"{x:.0f}"))
        plt.gca().xaxis.set_minor_locator(ticker.MultipleLocator(0.25))

        plt.tight_layout()
        plt.show()

    def export_to_excel(self, filename: str, cumulative_costs: np.ndarray | None = None) -> None:  # noqa: C901
        """Export simulation results and input parameters to a formatted Excel file.

        Args:
            filename (str): The name of the Excel file to be created.
            cumulative_costs: A list of lists, where each inner list represents a month and
                contains the profit/loss values for each simulation at that month.
        """
        if cumulative_costs is None:
            _, cumulative_costs = self.simulate_costs_over_time()

        wb = Workbook()
        ws = wb.active
        ws.title = "Rental Cost Summary"

        # Define styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )
        alignment = Alignment(horizontal="center", vertical="center")

        # Input Parameters section
        ws["A1"] = "Input Parameters"
        ws["A1"].font = title_font
        ws.merge_cells("A1:B1")
        ws["A1"].alignment = alignment

        headers = ["Parameter", "Value"]
        ws.append(headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        params = [
            ("Initial Monthly Rent", f"${self.initial_rent:,.2f}"),
            ("Lease Term", f"{self.lease_term} years"),
            ("Total Years", f"{self.total_years} years"),
            ("Monthly Utility Cost", f"${self.utility_cost:,.2f}"),
            ("Annual Renters Insurance", f"${self.renters_insurance:,.2f}"),
            ("Moving Cost", f"${self.moving_cost:,.2f}"),
            ("Mean Rent Increase Rate", f"{self.mean_rent_increase_rate:.2%}"),
            ("Rent Increase Volatility", f"{self.rent_increase_volatility:.2%}"),
            ("Mean Inflation Rate", f"{self.mean_inflation_rate:.2%}"),
            ("Inflation Volatility", f"{self.inflation_volatility:.2%}"),
            ("Broker Fee Rate", f"{self.broker_fee_rate:.2%}"),
            ("Move Probability", f"{self.move_probability:.2%}"),
            ("Rent Increase Move Threshold", f"{self.rent_increase_move_threshold:.2%}"),
            ("Number of Simulations", f"{self.simulations}"),
        ]

        for param in params:
            ws.append(param)

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Cost Statistics section
        start_row = ws.max_row + 2
        ws[f"A{start_row}"] = "Cost Statistics"
        ws[f"A{start_row}"].font = title_font
        ws.merge_cells(f"A{start_row}:B{start_row}")
        ws[f"A{start_row}"].alignment = alignment

        ws.append(["Statistic", "Value"])
        for cell in ws[start_row + 1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        cost_stats = self.get_cost_statistics(cumulative_costs)
        for stat, value in cost_stats.items():
            ws.append([stat.capitalize(), f"${value:,.2f}"])

        for row in ws.iter_rows(min_row=start_row + 2, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Costs Over Time section
        start_row = ws.max_row + 2
        ws[f"A{start_row}"] = "Costs Over Time"
        ws[f"A{start_row}"].font = title_font
        ws.merge_cells(f"A{start_row}:D{start_row}")
        ws[f"A{start_row}"].alignment = alignment

        time_headers = ["Month", "Average Cumulative Cost", "Lower 95% CI", "Upper 95% CI"]
        ws.append(time_headers)
        for cell in ws[start_row + 1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        months = list(range(1, (12 * self.total_years) + 1))
        avg_costs = cumulative_costs.mean(axis=1)
        ci_lower = [
            stats.t.ppf(0.025, len(month_costs) - 1) * (stats.tstd(month_costs) / np.sqrt(len(month_costs)))
            for month_costs in cumulative_costs
        ]
        ci_upper = [
            stats.t.ppf(0.975, len(month_costs) - 1) * (stats.tstd(month_costs) / np.sqrt(len(month_costs)))
            for month_costs in cumulative_costs
        ]

        for month, avg, lower, upper in zip(months, avg_costs, ci_lower, ci_upper, strict=True):
            ws.append([month, f"${avg:,.2f}", f"${avg+lower:,.2f}", f"${avg+upper:,.2f}"])

        for row in ws.iter_rows(min_row=start_row + 2, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        wb.save(filename)
        print(f"Data exported to '{filename}'")


# Example usage
if __name__ == "__main__":
    calculator = NYCRentalCostCalculator(
        initial_rent=3_500.0,
        lease_term=1.0,
        total_years=30,
        utility_cost=200.0,
        renters_insurance=300.0,
        moving_cost=2_000.0,
        mean_rent_increase_rate=0.03,
        rent_increase_volatility=0.01,
        mean_inflation_rate=0.02,
        inflation_volatility=0.005,
        broker_fee_rate=0.15,
        move_probability=0.1,
        rent_increase_move_threshold=0.1,
        simulations=5_000,
    )

    results = calculator.get_cost_statistics()
    print(f"Rental Cost Statistics: {results}")

    calculator.plot_costs_over_time()
