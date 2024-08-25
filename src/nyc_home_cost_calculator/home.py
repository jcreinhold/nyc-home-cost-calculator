from __future__ import annotations

from enum import Enum

import matplotlib.pyplot as plt
import numpy as np
from matplotlib import ticker
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy import stats
from tqdm.auto import tqdm


class FilingStatus(Enum):
    """Tax filing status categories for U.S. federal income tax.

    Attributes:
        SINGLE: Individual who is unmarried or legally separated.
        MARRIED_JOINT: Married couple filing a joint return.
        MARRIED_SEPARATE: Married individual filing separately.
    """

    SINGLE = "single"
    MARRIED_JOINT = "married_filing_jointly"
    MARRIED_SEPARATE = "married_filing_separately"


class NYCHomeCostCalculator:
    """A class to calculate and simulate the long-term costs and potential profits/losses of home ownership in NYC.

    This calculator takes into account various factors such as property appreciation,
    income changes, market volatility, taxes, and other associated costs of homeownership.
    It uses Monte Carlo simulation to project potential financial outcomes over time.
    """

    def __init__(
        self,
        home_price: float = 1_000_000.0,
        down_payment: float = 200_000.0,
        mortgage_rate: float = 0.05,
        loan_term: int = 30,
        initial_income: float = 150_000.0,
        initial_age: int = 30,
        hoa_fee: float = 500.0,
        insurance_rate: float = 0.005,
        maintenance_rate: float = 0.01,
        property_tax_rate: float = 0.01,
        mean_appreciation_rate: float = 0.03,
        appreciation_volatility: float = 0.05,
        mean_inflation_rate: float = 0.02,
        inflation_volatility: float = 0.01,
        mean_income_change_rate: float = 0.03,
        income_change_volatility: float = 0.03,
        retirement_contribution_rate: float = 0.15,
        filing_status: FilingStatus = FilingStatus.SINGLE,
        extra_payment: float = 0.0,
        extra_payment_start_year: int = 1,
        extra_payment_end_year: int = 30,
        purchase_closing_cost_rate: float = 0.03,
        sale_closing_cost_rate: float = 0.07,
        simulations: int = 5_000,
        rng: np.random.Generator | None = None,
    ):
        """Initialize the NYCHomeCostCalculator with home purchase and financial parameters.

        Args:
            home_price: The purchase price of the home. Defaults to $1,000,000.
            down_payment: The initial down payment amount. Defaults to $200,000.
            mortgage_rate: Annual mortgage interest rate (as a decimal). Defaults to 5%.
            loan_term: Length of the mortgage in years. Defaults to 30 (a 30-year mortgage).
            initial_income: Initial annual income of the homeowner. Defaults to $150,000.
            initial_age: Initial age of the homeowner. Defaults to 30.
            hoa_fee: Monthly homeowners association fee. Defaults to $500.
            insurance_rate: Annual insurance rate as a fraction of home value. Defaults to 0.5% of home value.
            maintenance_rate: Annual maintenance cost as a fraction of home value. Defaults to 1% of home value.
            property_tax_rate: Annual property tax rate. Defaults to 1% of home value.
            mean_appreciation_rate: Expected annual home appreciation rate. Defaults to 3% per year.
            appreciation_volatility: Volatility of the home appreciation rate. Defaults to 5%.
            mean_inflation_rate: Expected annual inflation rate. Defaults to 2% per year.
            inflation_volatility: Volatility of the inflation rate. Defaults to 1%.
            mean_income_change_rate: Expected annual rate of income change. Defaults to 3% per year.
            income_change_volatility: Volatility of annual income changes. Defaults to 3%.
            retirement_contribution_rate: Percentage of income contributed to retirement. Defaults to 15%.
            filing_status: Tax filing status category for U.S. federal income tax. Defaults to 'single'.
            extra_payment: An extra payment added toward principal. Defaults to $0.
            extra_payment_start_year: Year into loan term that extra payments are started. Defaults to the first year.
            extra_payment_end_year: Year into loan term that extra payments are ended. Defaults to year 30.
            purchase_closing_cost_rate: Closing costs for purchase as a percentage of home price. Defaults to 4%.
            sale_closing_cost_rate: Closing costs for sale as a percentage of sale price. Defaults to 7%.
            simulations: Number of Monte Carlo simulations to run. Defaults to 5,000.
            rng: Custom random number generator. If None, use default numpy RNG. Defaults to None.
        """
        # Store all input parameters as instance variables
        self.home_price = home_price
        self.down_payment = down_payment
        self.mortgage_rate = mortgage_rate
        self.loan_term = loan_term
        self.initial_income = initial_income
        self.hoa_fee = hoa_fee
        self.insurance_rate = insurance_rate
        self.maintenance_rate = maintenance_rate
        self.property_tax_rate = property_tax_rate
        self.mean_appreciation_rate = mean_appreciation_rate
        self.appreciation_volatility = appreciation_volatility
        self.mean_inflation_rate = mean_inflation_rate
        self.inflation_volatility = inflation_volatility
        self.mean_income_change_rate = mean_income_change_rate
        self.income_change_volatility = income_change_volatility
        self.retirement_contribution_rate = retirement_contribution_rate
        self.filing_status = filing_status
        self.extra_payment = extra_payment
        self.extra_payment_start_year = extra_payment_start_year
        self.extra_payment_end_year = extra_payment_end_year
        self.simulations = simulations
        self.rng = rng if rng is not None else np.random.default_rng()
        self.purchase_closing_cost_rate = purchase_closing_cost_rate
        self.sale_closing_cost_rate = sale_closing_cost_rate
        self.initial_age = initial_age
        self._current_year = 0

        # Define federal tax brackets (2023 rates)
        self.federal_brackets = self._get_federal_brackets()
        # Define New York State tax brackets (2023 rates)
        self.ny_state_brackets = self._get_ny_state_brackets()
        # Define New York City local tax rate (2023 rate)
        self.nyc_local_rate = 0.03876

        self.standard_deduction = self._get_standard_deduction()

    def _get_federal_brackets(self) -> list[tuple[float, float, float]]:
        if self.filing_status == FilingStatus.SINGLE:
            return [
                (0.0, 11_000.0, 0.10),
                (11_000.0, 44_725.0, 0.12),
                (44_725.0, 95_375.0, 0.22),
                (95_375.0, 182_100.0, 0.24),
                (182_100.0, 231_250.0, 0.32),
                (231_250.0, 578_125.0, 0.35),
                (578_125.0, float("inf"), 0.37),
            ]
        elif self.filing_status == FilingStatus.MARRIED_JOINT:
            return [
                (0.0, 22_000.0, 0.10),
                (22_000.0, 89_450.0, 0.12),
                (89_450.0, 190_750.0, 0.22),
                (190_750.0, 364_200.0, 0.24),
                (364_200.0, 462_500.0, 0.32),
                (462_500.0, 693_750.0, 0.35),
                (693_750.0, float("inf"), 0.37),
            ]
        elif self.filing_status == FilingStatus.MARRIED_SEPARATE:
            return [
                (0.0, 11_000.0, 0.10),
                (11_000.0, 44_725.0, 0.12),
                (44_725.0, 95_375.0, 0.22),
                (95_375.0, 182_100.0, 0.24),
                (182_100.0, 231_250.0, 0.32),
                (231_250.0, 346_875.0, 0.35),
                (346_875.0, float("inf"), 0.37),
            ]
        else:
            msg = "Invalid FilingStatus."
            raise ValueError(msg)

    def _get_ny_state_brackets(self) -> list[tuple[float, float, float]]:
        if self.filing_status == FilingStatus.SINGLE:
            return [
                (0.0, 8_500.0, 0.04),
                (8_500.0, 11_700.0, 0.045),
                (11_700.0, 13_900.0, 0.0525),
                (13_900.0, 80_650.0, 0.0585),
                (80_650.0, 215_400.0, 0.0625),
                (215_400.0, 1_077_550.0, 0.0685),
                (1_077_550.0, float("inf"), 0.0882),
            ]
        elif self.filing_status == FilingStatus.MARRIED_JOINT:
            return [
                (0.0, 17150.0, 0.04),
                (17_150.0, 23_600.0, 0.045),
                (23_600.0, 27_900.0, 0.0525),
                (27_900.0, 161_550.0, 0.0585),
                (161_550.0, 323_200.0, 0.0625),
                (323_200.0, 2_155_350.0, 0.0685),
                (2_155_350.0, float("inf"), 0.0882),
            ]
        elif self.filing_status == FilingStatus.MARRIED_SEPARATE:
            return [
                (0.0, 8_500.0, 0.04),
                (8_500.0, 11_800.0, 0.045),
                (11_800.0, 13_950.0, 0.0525),
                (13_950.0, 80_800.0, 0.0585),
                (80_800.0, 161_550.0, 0.0625),
                (161_550.0, 1_077_550.0, 0.0685),
                (1_077_550.0, float("inf"), 0.0882),
            ]
        else:
            msg = "Invalid FilingStatus."
            raise ValueError(msg)

    def _get_standard_deduction(self) -> float:
        if self.filing_status == FilingStatus.SINGLE:
            return 13_850.0
        elif self.filing_status == FilingStatus.MARRIED_JOINT:
            return 27_700.0
        elif self.filing_status == FilingStatus.MARRIED_SEPARATE:
            return 13_850.0
        else:
            msg = "Invalid FilingStatus."
            raise ValueError(msg)

    def get_salt_deduction_limit(self) -> float:
        """Get the SALT deduction limit based on filing status."""
        if self.filing_status == FilingStatus.MARRIED_SEPARATE:
            return 5_000.0
        return 10_000.0

    def get_mortgage_interest_deduction_limit(self) -> float:
        """Get the mortgage interest deduction limit based on filing status."""
        if self.filing_status == FilingStatus.MARRIED_SEPARATE:
            return 375_000.0 * self.mortgage_rate
        return 750_000.0 * self.mortgage_rate

    def calculate_tax_deduction(
        self,
        mortgage_interest: float,
        property_tax: float,
        federal_rate: float,
        state_rate: float,
        local_rate: float,
        retirement_contribution: float,
        annual_income: float,
    ) -> float:
        """Calculate tax deduction based on mortgage interest, property tax, retirement contributions, and tax rates.

        Args:
            mortgage_interest: Annual mortgage interest paid.
            property_tax: Annual property tax paid.
            federal_rate: Effective federal tax rate.
            state_rate: Effective state tax rate.
            local_rate: Effective local tax rate.
            retirement_contribution: Annual retirement contribution.
            annual_income: Annual income (used for calculating contribution limits).

        Returns:
            float: The calculated tax deduction amount.
        """
        # Get limits based on filing status
        salt_limit = self.get_salt_deduction_limit()
        mortgage_interest_limit = self.get_mortgage_interest_deduction_limit()

        # Calculate retirement contribution limit
        age = self._calculate_current_age()
        base_contribution_limit = 22_500  # 2023 limit for 401(k)
        catch_up_contribution = 7_500 if age >= 50 else 0  # 2023 catch-up contribution for age 50+
        retirement_contribution_limit = min(base_contribution_limit + catch_up_contribution, annual_income)

        # Cap the retirement contribution
        capped_retirement_contribution = min(retirement_contribution, retirement_contribution_limit)

        # Federal deduction calculation
        salt_deduction = min(salt_limit, property_tax)
        mortgage_interest_deduction = min(mortgage_interest, mortgage_interest_limit)
        federal_itemized_deduction = salt_deduction + mortgage_interest_deduction

        # Add capped retirement contribution to federal deduction (assuming traditional 401(k) or similar)
        federal_itemized_deduction += capped_retirement_contribution

        # State deduction calculation (New York doesn't limit SALT deductions)
        state_itemized_deduction = property_tax + mortgage_interest + capped_retirement_contribution

        federal_deduction = max(0, federal_itemized_deduction - self.standard_deduction) * federal_rate
        state_deduction = max(0, state_itemized_deduction - self.standard_deduction) * (state_rate + local_rate)

        return federal_deduction + state_deduction

    def _calculate_current_age(self) -> int:
        """Calculate the current age based on the simulation year."""
        return self.initial_age + self._current_year

    def calculate_tax(self, income: float, brackets: list[tuple[float, float, float]]) -> float:
        """Calculate the tax amount based on income and tax brackets.

        Args:
            income: The taxable income.
            brackets: List of tax brackets.
                Each bracket is a tuple of (lower_bound, upper_bound, tax_rate).

        Returns:
            The calculated tax amount.
        """
        tax = 0.0
        for lower, upper, rate in brackets:
            if income > lower:
                # Calculate tax for the portion of income in this bracket
                tax += (min(income, upper) - lower) * rate
            if income <= upper:
                # Break if we've reached the applicable bracket
                break
        return tax

    def generate_random_rates(self) -> tuple[float, float, float, float, float, float]:
        """Generate random rates for appreciation, inflation, income change, and tax adjustments.

        Returns:
            A tuple containing:
                - appreciation_rate: Random home appreciation rate
                - inflation_rate: Random inflation rate
                - income_change_rate: Random income change rate
                - federal_adjustment: Random federal tax rate adjustment
                - state_adjustment: Random state tax rate adjustment
                - local_adjustment: Random local tax rate adjustment
        """
        # Generate appreciation rate using t-distribution
        degrees_of_freedom = 3
        t_random = stats.t.rvs(degrees_of_freedom, random_state=self.rng.integers(1, 2**32 - 1))
        appreciation_rate = max(self.mean_appreciation_rate + self.appreciation_volatility * t_random, -0.5)

        # Generate inflation rate using normal distribution
        inflation_rate = self.rng.normal(self.mean_inflation_rate, self.inflation_volatility)

        # Generate income change rate using skew-normal distribution
        alpha = 2
        income_change_rate = stats.skewnorm.rvs(
            alpha,
            loc=self.mean_income_change_rate,
            scale=self.income_change_volatility,
            random_state=self.rng.integers(1, 2**32 - 1),
        )
        income_change_rate = max(income_change_rate, -0.5)  # Limit income loss to 50%

        # Generate random adjustments for tax rates
        federal_adjustment = self.rng.normal(0.0, 0.01)
        state_adjustment = self.rng.normal(0.0, 0.005)
        local_adjustment = self.rng.normal(0.0, 0.002)

        return (
            appreciation_rate,
            inflation_rate,
            income_change_rate,
            federal_adjustment,
            state_adjustment,
            local_adjustment,
        )

    def calculate_effective_tax_rates(
        self, income: float, federal_adj: float, state_adj: float, local_adj: float
    ) -> tuple[float, float, float]:
        """Calculate effective tax rates based on income and random adjustments.

        Args:
            income: The taxable income.
            federal_adj: Random adjustment to federal tax rate.
            state_adj: Random adjustment to state tax rate.
            local_adj: Random adjustment to local tax rate.

        Returns:
            Effective federal, state, and local tax rates.
        """
        federal_tax = self.calculate_tax(income, self.federal_brackets)
        state_tax = self.calculate_tax(income, self.ny_state_brackets)
        local_tax = income * self.nyc_local_rate

        # Calculate effective rates and apply random adjustments
        federal_rate = (federal_tax / income) + federal_adj
        state_rate = (state_tax / income) + state_adj
        local_rate = (local_tax / income) + local_adj

        # Ensure rates are non-negative
        return max(0.0, federal_rate), max(0.0, state_rate), max(0.0, local_rate)

    def calculate_monthly_payment(self, principal: float, annual_rate: float, months: int) -> float:
        """Calculate the monthly mortgage payment.

        Args:
            principal: The loan principal amount.
            annual_rate: Annual interest rate (as a decimal).
            months: The total number of monthly payments.

        Returns:
            The calculated monthly payment amount.
        """
        monthly_rate = annual_rate / 12.0
        return principal * (monthly_rate * (1.0 + monthly_rate) ** months) / ((1.0 + monthly_rate) ** months - 1.0)

    def simulate_costs_over_time(self) -> tuple[np.ndarray, np.ndarray]:
        """Simulate the costs and potential profits/losses of home ownership over time.

        This method runs Monte Carlo simulations for the entire loan term, taking into account
        various factors such as property appreciation, income changes, and market fluctuations.

        Returns:
            A tuple containing two NumPy arrays:
            1. An array of shape (simulations, loan_term * 12) containing monthly profit/loss values.
            2. An array of shape (simulations, loan_term * 12) containing cumulative profit/loss values.
        """
        # Calculate initial costs and loan details
        purchase_closing_costs = self.home_price * self.purchase_closing_cost_rate
        loan_amount = self.home_price - self.down_payment + purchase_closing_costs
        monthly_payment = self.calculate_monthly_payment(loan_amount, self.mortgage_rate, self.loan_term * 12)

        total_months = self.loan_term * 12
        monthly_costs = np.zeros((total_months, self.simulations))
        cumulative_costs = np.zeros((total_months, self.simulations))

        for sim in tqdm(range(self.simulations), desc="Running simulations", unit="sim"):
            # Initialize simulation variables
            home_value = self.home_price
            remaining_balance = loan_amount
            current_hoa_fee = self.hoa_fee
            current_income = self.initial_income
            cumulative_cost = purchase_closing_costs
            mortgage_paid_off = False
            self._current_year = 0

            for month in range(total_months):
                year = month // 12
                self._current_year = year

                # Generate random rates for this year (if it's a new year)
                if month % 12 == 0:
                    (
                        appreciation_rate,
                        inflation_rate,
                        income_change_rate,
                        fed_adj,
                        state_adj,
                        local_adj,
                    ) = self.generate_random_rates()

                    # Calculate retirement contribution (for tax purposes only)
                    annual_retirement_contribution = current_income * self.retirement_contribution_rate

                    # Calculate effective tax rates
                    federal_rate, state_rate, local_rate = self.calculate_effective_tax_rates(
                        current_income, fed_adj, state_adj, local_adj
                    )

                # Calculate monthly mortgage payment
                if not mortgage_paid_off:
                    monthly_interest = remaining_balance * (self.mortgage_rate / 12.0)
                    monthly_principal = min(monthly_payment - monthly_interest, remaining_balance)
                    remaining_balance -= monthly_principal

                    # Apply extra payment if within the specified period
                    if self.extra_payment_start_year <= year + 1 <= self.extra_payment_end_year:
                        extra_payment_this_month = min(self.extra_payment / 12.0, remaining_balance)
                        monthly_principal += extra_payment_this_month
                        remaining_balance -= extra_payment_this_month

                    if remaining_balance <= 0.0:
                        mortgage_paid_off = True
                        remaining_balance = 0.0

                # Calculate monthly costs
                property_tax = (home_value * self.property_tax_rate) / 12.0
                insurance = (home_value * self.insurance_rate) / 12.0
                maintenance = (home_value * self.maintenance_rate) / 12.0
                hoa = current_hoa_fee

                # Approximate monthly tax deduction (assuming even distribution throughout the year)
                tax_deduction = (
                    self.calculate_tax_deduction(
                        monthly_interest * 12.0,
                        property_tax * 12.0,
                        federal_rate,
                        state_rate,
                        local_rate,
                        annual_retirement_contribution,
                        current_income,
                    )
                    / 12.0
                )

                # Sum up total monthly cost
                monthly_cost = (
                    (monthly_payment if not mortgage_paid_off else 0.0)
                    + property_tax
                    + insurance
                    + maintenance
                    + hoa
                    - tax_deduction
                )
                cumulative_cost += monthly_cost

                # Calculate profit/loss for this month
                sale_closing_costs = home_value * self.sale_closing_cost_rate
                profit_loss = (
                    (home_value - self.home_price)
                    + (loan_amount - remaining_balance)
                    - cumulative_cost
                    - sale_closing_costs
                )

                monthly_costs[month, sim] = monthly_cost
                cumulative_costs[month, sim] = profit_loss

                # Update values for next month
                home_value *= (1.0 + appreciation_rate) ** (1.0 / 12.0)
                current_hoa_fee *= (1.0 + inflation_rate) ** (1.0 / 12.0)
                current_income *= (1.0 + income_change_rate) ** (1.0 / 12.0)

        return monthly_costs, cumulative_costs

    def get_cost_statistics(self, cumulative_costs: np.ndarray | None = None) -> dict[str, float]:
        """Calculate summary statistics of the simulated costs.

        This method runs the cost simulations and computes various statistical measures
        based on the final year's profit/loss values across all simulations.

        Args:
            cumulative_costs: A list of lists, where each inner list represents a month and
                contains the profit/loss values for each simulation at that month.

        Returns:
            A dictionary containing the following statistics:
                - mean: Average profit/loss
                - median: Median profit/loss
                - std_dev: Standard deviation of profit/loss
                - percentile_5: 5th percentile of profit/loss
                - percentile_95: 95th percentile of profit/loss
        """
        if cumulative_costs is None:
            _, cumulative_costs = self.simulate_costs_over_time()

        final_year_costs = cumulative_costs[-1]  # Get the profit/loss values for the final year

        return {
            "mean": final_year_costs.mean(),
            "median": np.quantile(final_year_costs, 0.5),
            "std_dev": np.std(final_year_costs, ddof=1),
            "percentile_5": np.quantile(final_year_costs, 0.05),
            "percentile_95": np.quantile(final_year_costs, 0.95),
        }

    def plot_costs_over_time(self, cumulative_costs: np.ndarray | None = None) -> None:
        """Plot the projected profit/loss over time with confidence intervals.

        This method generates a line plot showing the average profit/loss for each month
        of the loan term, along with a 95% confidence interval. It also includes a
        break-even line for reference.

        Args:
            cumulative_costs: A list of lists, where each inner list represents a month and
                contains the profit/loss values for each simulation at that month.
        """
        if cumulative_costs is None:
            _, cumulative_costs = self.simulate_costs_over_time()

        years = np.array(list(range(1, (12 * self.loan_term) + 1))) / 12
        avg_costs = cumulative_costs.mean(axis=1)

        # Calculate 95% confidence intervals
        ci_lower = [
            stats.t.ppf(0.025, len(year_costs) - 1) * (stats.tstd(year_costs) / np.sqrt(len(year_costs)))
            for year_costs in cumulative_costs
        ]
        ci_upper = [
            stats.t.ppf(0.975, len(year_costs) - 1) * (stats.tstd(year_costs) / np.sqrt(len(year_costs)))
            for year_costs in cumulative_costs
        ]

        lower_bound = [avg + ci_low for avg, ci_low in zip(avg_costs, ci_lower, strict=True)]
        upper_bound = [avg + ci_up for avg, ci_up in zip(avg_costs, ci_upper, strict=True)]

        # Create the plot
        plt.figure(figsize=(12, 6))
        plt.plot(years, avg_costs, label="Average Profit/Loss")
        plt.fill_between(years, lower_bound, upper_bound, alpha=0.2, label="95% Confidence Interval")
        plt.axhline(y=0, color="r", linestyle="--", label="Break-even")

        # Set plot labels and title
        plt.title("Projected Profit/Loss Over Time (Including Closing Costs)")
        plt.xlabel("Years")
        plt.ylabel("Profit/Loss ($)")
        plt.legend()
        plt.grid(True)

        # Format axes
        plt.gca().yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, p: f"${x:,.0f}"))
        plt.gca().xaxis.set_major_locator(ticker.MultipleLocator(1))
        plt.gca().xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, p: f"{x:.0f}"))
        plt.gca().xaxis.set_minor_locator(ticker.MultipleLocator(0.25))

        plt.tight_layout()
        plt.show()

    def export_to_excel(self, filename: str, cumulative_costs: np.ndarray | None = None) -> None:  # noqa: C901
        """Export simulation results and input parameters to an Excel file.

        This method creates a formatted Excel workbook with three main sections:
        1. Input Parameters
        2. Cost Statistics
        3. Profit/Loss Over Time

        Args:
            filename: The name of the Excel file to be created.
            cumulative_costs: A list of lists, where each inner list represents a month and
                contains the profit/loss values for each simulation at that month.
        """
        if cumulative_costs is None:
            _, cumulative_costs = self.simulate_costs_over_time()

        wb = Workbook()
        ws = wb.active
        ws.title = "Home Ownership Cost Summary"

        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )
        alignment = Alignment(horizontal="center", vertical="center")

        # Input Parameters section
        ws["A1"] = "Input Parameters"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")
        ws["A1"].alignment = alignment

        headers = ["Parameter", "Value"]
        ws.append(headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        params = [
            ("Home Price", f"${self.home_price:,.2f}"),
            ("Down Payment", f"${self.down_payment:,.2f}"),
            ("Mortgage Rate", f"{self.mortgage_rate:.2%}"),
            ("Loan Term", f"{self.loan_term} years"),
            ("Initial Annual Income", f"${self.initial_income:,.2f}"),
            ("Monthly HOA Fee", f"${self.hoa_fee:,.2f}"),
            ("Insurance Rate", f"{self.insurance_rate:.2%}"),
            ("Maintenance Rate", f"{self.maintenance_rate:.2%}"),
            ("Property Tax Rate", f"{self.property_tax_rate:.2%}"),
            ("Mean Appreciation Rate", f"{self.mean_appreciation_rate:.2%}"),
            ("Appreciation Volatility", f"{self.appreciation_volatility:.2%}"),
            ("Mean Inflation Rate", f"{self.mean_inflation_rate:.2%}"),
            ("Inflation Volatility", f"{self.inflation_volatility:.2%}"),
            ("Mean Income Change Rate", f"{self.mean_income_change_rate:.2%}"),
            ("Income Change Volatility", f"{self.income_change_volatility:.2%}"),
            ("Retirement Contribution Rate", f"{self.retirement_contribution_rate:.2%}"),
            ("Filing Status", f"{self.filing_status.value}"),
            ("Extra Payment", f"${self.extra_payment:,.2f}"),
            ("Extra Payment Start Year", f"{self.extra_payment_start_year}"),
            ("Extra Payment End Year", f"{self.extra_payment_end_year}"),
            ("Purchase Closing Cost Rate", f"{self.purchase_closing_cost_rate:.2%}"),
            ("Sale Closing Cost Rate", f"{self.sale_closing_cost_rate:.2%}"),
            ("Number of Simulations", f"{self.simulations}"),
        ]

        for param in params:
            ws.append(param)

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Cost Statistics section
        start_row = ws.max_row + 2
        ws[f"A{start_row}"] = "Cost Statistics"
        ws[f"A{start_row}"].font = Font(bold=True, size=14)
        ws.merge_cells(f"A{start_row}:B{start_row}")
        ws[f"A{start_row}"].alignment = alignment

        ws.append(["Statistic", "Value"])
        for cell in ws[start_row + 1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        cost_stats = self.get_cost_statistics(cumulative_costs)
        for stat, value in cost_stats.items():
            ws.append([stat.capitalize(), f"${value:,.2f}"])

        for row in ws.iter_rows(min_row=start_row + 2, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Profit/Loss Over Time section
        start_row = ws.max_row + 2
        ws[f"A{start_row}"] = "Profit/Loss Over Time"
        ws[f"A{start_row}"].font = Font(bold=True, size=14)
        ws.merge_cells(f"A{start_row}:D{start_row}")
        ws[f"A{start_row}"].alignment = alignment

        time_headers = ["Month", "Average Profit/Loss", "Lower 95% CI", "Upper 95% CI"]
        ws.append(time_headers)
        for cell in ws[start_row + 1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        months = list(range(1, (12 * self.loan_term) + 1))
        avg_costs = cumulative_costs.mean(axis=1)
        ci_lower = [
            stats.t.ppf(0.025, len(month_costs) - 1) * (stats.tstd(month_costs) / np.sqrt(len(month_costs)))
            for month_costs in cumulative_costs
        ]
        ci_upper = [
            stats.t.ppf(0.975, len(month_costs) - 1) * (stats.tstd(month_costs) / np.sqrt(len(month_costs)))
            for month_costs in cumulative_costs
        ]

        for month, avg, lower, upper in zip(months, avg_costs, ci_lower, ci_upper, strict=True):
            ws.append([month, f"${avg:,.2f}", f"${avg+lower:,.2f}", f"${avg+upper:,.2f}"])

        for row in ws.iter_rows(min_row=start_row + 2, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        wb.save(filename)
        print(f"Data exported to '{filename}'")


# Example usage
if __name__ == "__main__":
    custom_rng = np.random.default_rng(42)

    calculator = NYCHomeCostCalculator(
        home_price=1_000_000,
        down_payment=200_000,
        mortgage_rate=0.03,
        loan_term=30,
        initial_income=150_000,
        hoa_fee=500,
        insurance_rate=0.005,
        maintenance_rate=0.01,
        property_tax_rate=0.01,
        mean_appreciation_rate=0.03,
        appreciation_volatility=0.05,
        mean_inflation_rate=0.02,
        inflation_volatility=0.01,
        mean_income_change_rate=0.02,
        income_change_volatility=0.03,
        retirement_contribution_rate=0.15,
        filing_status=FilingStatus.SINGLE,
        simulations=10_000,
        rng=custom_rng,
        purchase_closing_cost_rate=0.04,
        sale_closing_cost_rate=0.08,
    )

    # Calculate and print cost statistics
    results = calculator.get_cost_statistics()
    print(f"Cost Statistics: {results}")

    # Plot costs over time
    calculator.plot_costs_over_time()

    # Export data to Excel
    calculator.export_to_excel("nyc-home-ownership-summary.xlsx")
