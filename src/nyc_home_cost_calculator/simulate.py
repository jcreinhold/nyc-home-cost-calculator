"""This module provides classes and functions for simulating home costs in NYC."""

from __future__ import annotations

import datetime
import logging
from dataclasses import asdict, dataclass, field
from typing import TYPE_CHECKING, Any, Protocol, TypeVar

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import ticker
from matplotlib.colors import to_rgba
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from nyc_home_cost_calculator.utils import calculate_confidence_intervals

if TYPE_CHECKING:  # pragma: no cover
    from collections.abc import Callable

    from matplotlib.axes import Axes

    T = TypeVar("T")

logger = logging.getLogger(__name__)


@dataclass
class SimulationResults:
    """Represents the results of various financial simulations."""

    # Common fields
    monthly_costs: np.ndarray
    profit_loss: np.ndarray
    total_years: float | int
    simulations: int

    # Home ownership specific
    home_values: np.ndarray | None = None
    remaining_mortgage_balance: np.ndarray | None = None
    property_taxes: np.ndarray | None = None
    insurance_costs: np.ndarray | None = None
    maintenance_costs: np.ndarray | None = None

    # Investment specific
    portfolio_values: np.ndarray | None = None
    investment_returns: np.ndarray | None = None

    # Career and life event specific
    personal_income: np.ndarray | None = None
    monthly_income: np.ndarray | None = None
    promotions: np.ndarray | None = None
    demotions: np.ndarray | None = None
    layoffs: np.ndarray | None = None
    marital_status: np.ndarray | None = None
    partner_income: np.ndarray | None = None
    household_income: np.ndarray | None = None

    # Tax related
    tax_deductions: np.ndarray | None = None
    federal_effective_tax_rate: np.ndarray | None = None
    state_effective_tax_rate: np.ndarray | None = None
    local_effective_tax_rate: np.ndarray | None = None
    after_tax_income: np.ndarray | None = None

    # Misc financial
    cumulative_costs: np.ndarray | None = None
    cumulative_savings: np.ndarray | None = None

    # Additional fields for flexibility
    extra: dict = field(default_factory=dict)

    # Internal field for recording input parameters
    _input_parameters: list[tuple[str, Any]] = field(default_factory=list)

    def __repr__(self):
        """Provide a concise representation of the SimulationResults."""
        fields = [
            f"{k}={v.shape if isinstance(v, np.ndarray) else v}"
            for k, v in asdict(self).items()
            if v is not None and not k.startswith("_") and k != "extra"
        ]
        return f"SimulationResults({', '.join(fields)})"

    def get_percentiles(self, field: str, percentiles: list[float]) -> np.ndarray:
        """Calculate percentiles for a given field across all simulations.

        Args:
            field: The name of the field to calculate percentiles for.
            percentiles: List of percentiles to calculate (e.g., [0.05, 0.5, 0.95]).

        Returns:
            Array of percentile values for each time step.
        """
        data = getattr(self, field)
        if data is None:
            msg = f"Field '{field}' not found in simulation results."
            raise ValueError(msg)
        return np.percentile(data, percentiles, axis=1)

    def get_mean(self, field: str) -> np.ndarray:
        """Calculate the mean for a given field across all simulations.

        Args:
            field: The name of the field to calculate the mean for.

        Returns:
            Array of mean values for each time step.
        """
        data = getattr(self, field)
        if data is None:
            msg = f"Field '{field}' not found in simulation results."
            raise ValueError(msg)
        return np.mean(data, axis=1)

    def get_std_dev(self, field: str) -> np.ndarray:
        """Calculate the standard deviation for a given field across all simulations.

        Args:
            field: The name of the field to calculate the standard deviation for.

        Returns:
            Array of standard deviation values for each time step.
        """
        data = getattr(self, field)
        if data is None:
            msg = f"Field '{field}' not found in simulation results."
            raise ValueError(msg)
        return np.std(data, axis=1)

    def __add__(self, other: SimulationResults) -> SimulationResults:
        """Add two SimulationResults objects together.

        Args:
            other: The SimulationResults object to add.

        Returns:
            The combined SimulationResults object.
        """
        return self._combine(other, np.add)

    def __sub__(self, other: SimulationResults) -> SimulationResults:
        """Subtract two SimulationResults objects.

        Args:
            other: The SimulationResults object to subtract.

        Returns:
            The subtracted SimulationResults object.
        """
        return self._combine(other, np.subtract)

    def __mul__(self, other: SimulationResults) -> SimulationResults:
        """Multiply two SimulationResults objects together.

        Args:
            other: The SimulationResults object to multiply.

        Returns:
            The multiplied SimulationResults object.
        """
        return self._combine(other, np.multiply)

    def __truediv__(self, other: SimulationResults) -> SimulationResults:
        """Divide two SimulationResults objects element-wise.

        Args:
            other: The SimulationResults object to divide.

        Returns:
            The divided SimulationResults object.
        """
        return self._combine(other, np.divide)

    def _combine(self, other: SimulationResults, operation: Callable) -> SimulationResults:
        if not isinstance(other, SimulationResults):
            msg = f"Unsupported operand type for SimulationResults: {type(other)}"
            raise TypeError(msg)

        def _combine_arrays(a: T, b: T) -> T | dict | int | float | None:
            if a is None:
                return b
            if b is None:
                return a
            if isinstance(a, int | float) and isinstance(b, int | float) and a == b:
                return a
            if isinstance(a, np.ndarray) and isinstance(b, np.ndarray):
                return operation(a, b)
            if isinstance(a, dict) and isinstance(b, dict):
                return {**a, **b}
            return None

        combined_dict: dict[str, np.ndarray | dict | int | float | None] = {}
        for _field in self.__dataclass_fields__:
            combined_dict[_field] = _combine_arrays(getattr(self, _field), getattr(other, _field))

        return SimulationResults(**combined_dict)  # type: ignore[arg-type]

    @classmethod
    def zeros_like(cls, other: SimulationResults) -> SimulationResults:
        """Create a new SimulationResults instance with zero-filled arrays like another instance."""
        zero_dict: dict[str, Any] = {}
        for _field, value in asdict(other).items():
            if isinstance(value, np.ndarray):
                zero_dict[_field] = np.zeros_like(value)
            elif _field == "extra":
                zero_dict[_field] = {}
            else:
                zero_dict[_field] = None
        return cls(**zero_dict)

    def to_dataframe(self, *, aggregator: _AggregatorProtocol = np.mean) -> pd.DataFrame:
        """Convert the SimulationResults object to a pandas DataFrame.

        Args:
            aggregator: The function used to aggregate the data for each time step.
                Defaults to np.mean.

        Returns:
            A pandas DataFrame representing the SimulationResults object.
        """

        def process_value(key: str, value: Any) -> pd.Series | pd.DataFrame:
            if isinstance(value, np.ndarray):
                if value.ndim == 2:
                    return pd.Series(aggregator(value, axis=1), name=key)
                return pd.Series(value, name=key)
            if isinstance(value, dict):
                return pd.concat([process_value(f"{key}_{k}", v) for k, v in value.items()], axis=1)
            return pd.Series([value] * len(self.monthly_costs), name=key)

        data = pd.concat(
            [process_value(field, value) for field, value in self.__dict__.items() if not field.startswith("_")], axis=1
        )

        start_date = datetime.datetime.now(datetime.timezone.utc).replace(day=1)
        date_index = pd.date_range(start=start_date, periods=len(self.monthly_costs), freq="M")

        return pd.DataFrame(data, index=date_index)

    def get_cost_statistics(self) -> dict[str, float]:
        """Calculate summary statistics of the simulated costs.

        This method runs the cost simulations and computes various statistical measures
        based on the final year's profit/loss values across all simulations.

        Returns:
            A dictionary containing the following statistics:
                - mean: Average profit/loss
                - median: Median profit/loss
                - std_dev: Standard deviation of profit/loss
                - percentile_5: 5th percentile of profit/loss
                - percentile_95: 95th percentile of profit/loss
        """
        profit_loss = self.profit_loss

        if profit_loss is None:
            msg = "Could not find 'profit_loss'."
            raise ValueError(msg)

        final_year_costs = profit_loss[-1]

        return {
            "mean": final_year_costs.mean(),
            "median": np.quantile(final_year_costs, 0.5),
            "std_dev": np.std(final_year_costs, ddof=1),
            "percentile_5": np.quantile(final_year_costs, 0.05),
            "percentile_95": np.quantile(final_year_costs, 0.95),
        }

    def plot(
        self,
        ax: Axes | None = None,
        *,
        figsize: tuple[int, int] = (12, 6),
        title: str = "Projected Profit Over Time",
        ylabel: str = "Profit ($)",
        label: str = "Average Profit",
        ci: float = 0.95,
        quantile_range: tuple[float, float] | None = None,
        color: str = "C0",
    ) -> Axes:
        """Plot the projected profit/loss over time with confidence intervals and optional percentile range.

        This method generates a line plot showing the average profit/loss for each month
        of the loan term, along with a confidence interval and an optional percentile range.

        Args:
            ax: Optional matplotlib Axes object to plot on. If not provided, a new figure and axes will be created.
            figsize: A tuple specifying the figure size in inches (width, height). Only used if ax is None.
            title: The title of the plot.
            ylabel: The label for the y-axis.
            label: The label for the average cost line.
            ci: The confidence interval to plot (default: 0.95 for 95% CI).
            quantile_range: An optional tuple of (lower, upper) quantiles to plot (e.g., (0.05, 0.95) for 90% range).
            color: The base color to use for the plot (default: 'C0', which is usually blue).

        Returns:
            The matplotlib Axes object containing the plot.
        """
        profit_loss = self.profit_loss

        if profit_loss is None:
            msg = "Could not find profit/loss data."
            raise ValueError(msg)

        years = np.arange(1, int((12 * self.total_years) + 1)) / 12
        avg_costs, lower_bound, upper_bound = calculate_confidence_intervals(profit_loss, confidence_level=ci)

        if ax is None:  # pragma: no cover
            _, ax = plt.subplots(figsize=figsize)

        # Convert color to RGBA
        base_color = to_rgba(color)

        # Plot average line
        ax.plot(years, avg_costs, color=base_color, label=label)

        # Plot confidence intervals
        ci_color = (*base_color[:3], 0.3)
        ax.fill_between(years, lower_bound, upper_bound, color=ci_color, label=f"{100.0 * ci:.0f}% Confidence Interval")

        if quantile_range:
            lower_quantile, upper_quantile = quantile_range
            lower_bound = np.quantile(profit_loss, lower_quantile, axis=1)
            upper_bound = np.quantile(profit_loss, upper_quantile, axis=1)
            quantile_color = (*base_color[:3], 0.1)
            ax.fill_between(
                years,
                lower_bound,
                upper_bound,
                color=quantile_color,
                label=f"{100.0 * (upper_quantile - lower_quantile):.0f}% Range",
            )

        ax.set_title(title)
        ax.set_xlabel("Years")
        ax.set_ylabel(ylabel)
        ax.legend()
        ax.grid(visible=True)

        ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
        ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
        ax.xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"{x:.0f}"))
        ax.xaxis.set_minor_locator(ticker.MultipleLocator(0.25))

        plt.tight_layout()
        return ax

    def export_to_excel(self, filename: str) -> None:  # noqa: C901, PLR0912, PLR0914, PLR0915
        """Export simulation results and input parameters to an Excel file.

        This method creates a formatted Excel workbook with three main sections:
        1. Input Parameters
        2. Cost Statistics
        3. Profit/Loss Over Time

        Args:
            filename: The name of the Excel file to be created.
        """
        profit_loss = self.profit_loss

        if profit_loss is None:
            msg = "Could not find 'profit_loss'."
            raise ValueError(msg)

        wb = Workbook()
        ws = wb.active
        ws.title = "Cost Summary"

        # Define styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )
        alignment = Alignment(horizontal="center", vertical="center")

        # Input Parameters section
        ws["A1"] = "Input Parameters"
        ws["A1"].font = title_font
        ws.merge_cells("A1:B1")
        ws["A1"].alignment = alignment

        headers = ["Parameter", "Value"]
        ws.append(headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        params = self._input_parameters
        for param in params:
            ws.append(param)

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Cost Statistics section
        start_row = ws.max_row + 2
        ws[f"A{start_row}"] = "Cost Statistics"
        ws[f"A{start_row}"].font = title_font
        ws.merge_cells(f"A{start_row}:B{start_row}")
        ws[f"A{start_row}"].alignment = alignment

        ws.append(["Statistic", "Value"])
        for cell in ws[start_row + 1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        cost_stats = self.get_cost_statistics()
        for stat, value in cost_stats.items():
            ws.append([stat.capitalize(), f"${value:,.2f}"])

        for row in ws.iter_rows(min_row=start_row + 2, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Costs Over Time section
        start_row = ws.max_row + 2
        ws[f"A{start_row}"] = "Costs Over Time"
        ws[f"A{start_row}"].font = title_font
        ws.merge_cells(f"A{start_row}:D{start_row}")
        ws[f"A{start_row}"].alignment = alignment

        time_headers = ["Month", "Average Cost", "Lower 95% CI", "Upper 95% CI"]
        ws.append(time_headers)
        for cell in ws[start_row + 1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        months = range(1, int((12 * self.total_years) + 1))
        avg_costs, ci_lower, ci_upper = calculate_confidence_intervals(profit_loss)

        for month, avg, lower, upper in zip(months, avg_costs, ci_lower, ci_upper, strict=True):
            ws.append([month, f"${avg:,.2f}", f"${avg + lower:,.2f}", f"${avg + upper:,.2f}"])

        for row in ws.iter_rows(min_row=start_row + 2, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except (AttributeError, TypeError) as e:
                    logger.debug("Error: %s", e)
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        wb.save(filename)
        logger.info("Data exported to '%s'", filename)


class SimulationEngine:
    """Represents a simulation engine for calculating home costs in NYC."""

    def __init__(self, total_months: int, simulations: int, rng: np.random.Generator | None = None):
        """Initialize the SimulationEngine object.

        Args:
            total_months: The total number of months for the simulation.
            simulations: The number of simulations to run.
            rng: The random number generator to use.
        """
        self.total_months = total_months
        self.simulations = simulations
        self.rng = rng or np.random.default_rng()

    def run_simulation(self, simulate_vectorized: Callable[[np.ndarray], SimulationResults]) -> SimulationResults:
        """Run the simulation using the provided vectorized simulation function.

        Args:
            simulate_vectorized: A vectorized simulation function that takes an array of months as input and returns
                the monthly costs and cumulative costs.

        Returns:
            The results of the simulation, including the monthly costs and profit/loss.
        """
        months = np.arange(self.total_months)[:, np.newaxis]
        months_matrix = np.tile(months, (1, self.simulations))
        return simulate_vectorized(months_matrix)


class _AggregatorProtocol(Protocol):
    def __call__(self, a: np.ndarray, axis: int, *args: Any, **kwargs: Any) -> np.ndarray: ...
